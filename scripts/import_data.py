import openpyxl
import os
import json
import re

def is_generic_dosage_form_list(text):
    if not text:
        return True
    
    text_clean = text.lower().strip()
    generic_terms = [
        'solid orals', 'liquid orals', 'topicals', 'ophthalmic', 'otic', 
        'otd/mups', 'otd /mups', 'coatings', 'pellets', 'parenterals', 
        'inhalation', 'excipient', 'food&beverage', 'dietary supplements',
        'feed', 'body care', 'sun care', 'hair care', 'make up', 'body&hair care'
    ]
    
    parts = re.split(r'[,;&/]| and |\b&\b', text_clean)
    parts = [p.strip() for p in parts if p.strip()]
    
    if not parts:
        return True
        
    for part in parts:
        matched = False
        for gt in generic_terms:
            if part in gt or gt in part or re.sub(r'\s+', '', part) in re.sub(r'\s+', '', gt):
                matched = True
                break
        if not matched:
            return False
            
    return True

# Define unmapped or mapped partner ID resolutions
def get_partner_id(supplier_name):
    name = str(supplier_name).strip().upper()
    if '3V SIGMA' in name:
        return '3v-sigma'
    if 'UNIPROMA' in name:
        return 'uniproma'
    if 'BIOSYNTIA' in name or 'BIOSYNTHIA' in name:
        return 'biosynthia'
    if 'NEWMAN' in name:
        return 'newman'
    if 'NHU' in name:
        return 'nhu'
    if 'NANOGEN' in name:
        return 'nanogen'
    if 'MFCI' in name:
        return 'mfci'
    if 'SPEC-CHEM' in name or 'SPEC CHEM' in name or 'SPECCHEM' in name:
        return 'sc-group'
    if 'TINCI' in name:
        return 'tinci'
    if 'LINGEBA' in name:
        return 'lingeba'
    if 'XINFU' in name:
        return 'xinfu'
    if 'HOWTIAN' in name or 'SO PURE' in name or 'SOPURE' in name or 'HAOTIAN' in name:
        return 'howtian'
    if 'NEXTHIA' in name:
        return 'nexthia'
    if 'SABO' in name:
        return 'sabo'
    if 'AKOTT' in name:
        return 'akott'
    if 'ALULA' in name:
        return 'alula'
    if 'SALVONA' in name:
        return 'salvona'
    if 'ZOCHEM' in name:
        return 'zochem'
    if 'OLVEA' in name:
        return 'olvea'
    if 'KIYU' in name:
        return 'kiyu'
    if 'KCI' in name or 'SAMYANG' in name:
        return 'kci'
    if 'AURORA' in name:
        return 'yantai'
    if 'XINYU' in name:
        return 'xinyu'
    if 'CHT' in name:
        return 'cht'
    if 'PELLETS' in name:
        return 'pellets'
    if 'LIPID CHEM' in name or 'LIPIDCHEM' in name or 'LIPICHEM' in name:
        return 'lipidchem'
    if 'ORIEN' in name:
        return 'orien'
    if 'SATORIA' in name:
        return 'satoria'
    if 'BALT' in name:
        return 'baltmilk'
    if 'TATE' in name:
        return 'tate-lyle'
    if 'JIANYUAN' in name:
        return 'jianyuan'
    if 'CYGNUS' in name:
        return 'cygnus'
    if 'DEEPAK' in name:
        return 'deepak-cellulose'
    if 'HUAKANG' in name:
        return 'huakang'
    if 'KLK' in name:
        return 'klk-oleo'
    if 'LACTO' in name:
        return 'lactopharm'
    if 'NUCELL' in name:
        return 'nucell'
    if 'PRAYON' in name:
        return 'prayon'
    if 'RECKON' in name:
        return 'reckon'
    if 'RUITAI' in name:
        return 'rutocel'
    if 'SMS' in name:
        return 'sms'
    if 'STANDARD' in name:
        return 'standard-chem'
    if 'SUN FLOWER' in name or 'SUNFLOWER' in name:
        return 'sun-flower'
    if 'SUN CHEMICAL' in name or 'SUNCHEMICAL' in name:
        return 'sun-chemical'
    if 'TIANLI' in name:
        return 'tianli'
    if 'ESSENTIAL' in name:
        return 'essential'
    if 'SILVATEAM' in name:
        return 'silvateam'
    if 'TNJ' in name:
        return 'tnj'
    
    clean_id = ''.join(c if c.isalnum() else '-' for c in name.lower())
    clean_id = '-'.join(filter(None, clean_id.split('-')))
    return clean_id

# Clean name mapping
def get_clean_partner_name(supplier_name, default_id):
    name = str(supplier_name).strip()
    # Normalize clean names for marquee/verticals
    clean_names = {
        '3v-sigma': '3V Sigma USA',
        'uniproma': 'Uniproma',
        'biosynthia': 'Biosyntia',
        'newman': 'Newman',
        'nhu': 'NHU',
        'nanogen': 'Nanogen',
        'mfci': 'MFCI',
        'sc-group': 'SC Group (SpecChem)',
        'tinci': 'Tinci',
        'lingeba': 'Lingeba',
        'xinfu': 'Xinfu',
        'howtian': 'Howtian / So Pure',
        'nexthia': 'Nexthia',
        'sabo': 'Sabo',
        'akott': 'Akott',
        'alula': 'Alula',
        'salvona': 'Salvona',
        'zochem': 'Zochem',
        'olvea': 'Olvea',
        'kiyu': 'Kiyu',
        'kci': 'KCI (Samyang)',
        'yantai': 'Yantai / Aurora Chem',
        'xinyu': 'Xinyu',
        'cht': 'CHT',
        'pellets': 'Pellets',
        'lipidchem': 'Lipidchem',
        'orien': 'Orien',
        'satoria': 'Satoria Agro',
        'baltmilk': 'BaltMilk',
        'tate-lyle': 'Tate & Lyle',
        'jianyuan': 'Jianyuan International',
        'cygnus': 'Cygnus Polyols',
        'deepak-cellulose': 'Deepak Cellulose',
        'huakang': 'Huakang',
        'klk-oleo': 'KLK Oleo',
        'lactopharm': 'LactoPharm',
        'nucell': 'NUCELL',
        'prayon': 'Prayon',
        'reckon': 'Reckon Organics',
        'rutocel': 'RUTOCEL',
        'sms': 'SMS Corporation',
        'standard-chem': 'Standard Chem',
        'sun-flower': 'Huzhou Sunflower',
        'sun-chemical': 'Sun Chemical',
        'tianli': 'Tianli',
        'essential': 'Essential Flavours',
        'silvateam': 'Silvateam',
        'tnj': 'TNJ'
    }
    return clean_names.get(default_id, name)

# Handcrafted specialties and bios
existing_partners = {
    "reckon": {"specialty": "Ascorbic acid & vitamin C derivatives", "about": "India GMP Certified company and leading manufacturer of Ascorbic Acid (Vitamin-C) and other excipients offers quality product and services."},
    "prayon": {"specialty": "Phosphate-based ingredients", "about": "Prayon is a global leader in phosphate chemistry. Headquartered in Belgium, the company manufactures a wide range of phosphate-based products for food, pharmaceutical, and industrial applications, delivering high-performance chemical solutions."},
    "standard-chem": {"specialty": "Sodium stearyl fumarate lubricant", "about": "Standard chem is one of the biggest and famous drug as well as chemical company in Taiwan. It strictly works with the philosophy of Sincerity, Honesty, Experience, Innovation and Public Service."},
    "rutocel": {"specialty": "HPMC, ethyl cellulose & HPC", "about": "Taian Ruitai Cellulose Co., Ltd. is a leading non-ionic cellulose manufacturer established in 1988. The company specializes in the research, production, and export of cellulose ethers including MC, HPMC, HPC, EC, HEC and film-coating products, widely used in pharmaceuticals, food, cosmetics, and coatings."},
    "tianli": {"specialty": "Mannitol & sorbitol polyols", "about": "It is the largest manufacturers of Polyols in Asia. Tianli belongs to Shandong Lianmeng Chemical Group. The company has sound management systems and certifications ie GMP, ISO, Kosher, HACCP, etc. The products are being exported to USA, Europe, South America, Asia, etc."},
    "cygnus": {"specialty": "Mannitol & lactose formulations", "about": "It is specialized in formulation solutions to aid the improvement of final pharma and food products. Audits, preparation of strict specifications, revisions, monitoring, and updating quality parameters are our routine jobs to make a strong alliance between Cygnus, our suppliers, and customers."},
    "lactopharm": {"specialty": "Lactose for tableting & capsules", "about": "It is a company specialized in milk derivatives, including milk protein, lactose, and other ingredients extracted from milk."},
    "sun-chemical": {"specialty": "Iron oxide colorants", "about": "Together with DIC, Sun Chemical has annual sales of more than $8.5 billion. Over 22,000 employees work every day to meet the needs of customers by improving performance on the essentials of business."},
    "nucell": {"specialty": "Functional pharma & nutraceutical ingredients", "about": "Nucell Pharma is a U.S.-based, quality-driven company providing functional ingredients for pharmaceutical, food, and cosmetic applications. They offer high-performance ingredients that help formulators enhance the quality of their finished products. Committed to cGMP standards, their Quality Assurance programs ensure every production batch complies with IPEC guidelines."},
    "klk-oleo": {"specialty": "Oleochemicals, fatty acids & specialty esters", "about": "KLK OLEO is a global leader in oleochemicals, committed to delivering sustainable and high-quality solutions. As the oleochemicals division of Kuala Lumpur Kepong Berhad (KLK), KLK OLEO operates integrated complexes in Malaysia, Indonesia, China, and Europe, producing a wide range of products from renewable raw materials including fatty acids, glycerine, fatty alcohols, esters, and specialty ingredients."},
    "sun-flower": {"specialty": "PVP, NVP & pharmaceutical polymers", "about": "Huzhou Sunflower Pharmaceutical Co., Ltd. is an ISO 9001 and ISO 14001 certified company with Europe DMF/US DMF dossier registrations and c-GMP approval. The company is a leading manufacturer of PVP and NVP series products, ranking among the largest producers in China and third globally, with a production capacity of over 8,000 MT per year."},
    "sms": {"specialty": "Modified tapioca starch for pharma & food", "about": "SMS Corporation Co., Ltd. is a global manufacturer of modified tapioca starch, supplying various industries including food, healthcare, personal care, and bioplastics. The company operates three manufacturing facilities in Thailand with a total production capacity of 400,000 tons per year."},
    "huakang": {"specialty": "Polyols: xylitol, mannitol, sorbitol & erythritol", "about": "Founded in 1962, Huakang is a global leader in polyols and plant-based ingredients, operating in more than 80 countries. As a national high-tech enterprise, the company specializes in the research, production, and sales of functional polyols, starch sugars, and health food ingredients, and is a major global producer of xylitol, mannitol, and maltitol."},
    "deepak-cellulose": {"specialty": "Cellulose ethers & fatty acid excipients", "about": "Deepak Cellulose Pvt. Ltd. is a leading Indian manufacturer of Cellulose Ether and Fatty Acid, supplying high-quality excipients to pharmaceutical, nutraceutical, and cosmetic industries. Established in 2004 and ISO 9001:2015 certified, the company has over 15 years of experience providing reliable excipients to the pharmaceutical sector."},
    "uniproma": {"specialty": "Personal care ingredients", "about": "Uniproma Chemical is a trusted partner in delivering innovative, high-performance solutions for cosmetics and personal care. Combining sustainable advancements in green chemistry with cutting-edge research, they specialize in eco-friendly raw materials, sunscreens, and skin-brightening actives."},
    "biosynthia": {"specialty": "Personal care ingredients", "about": "Biosyntia is a leading biotechnology company focused on developing green, bio-based ingredients for personal care and nutrition. Through advanced fermentation processes, they replace fossil-based chemicals with sustainable, natural alternatives."},
    "tnj": {"specialty": "Personal care ingredients", "about": "Hefei TNJ Chemical is a leading global manufacturer and supplier of fine chemicals, pharmaceutical excipients, and personal care ingredients, serving clients in over 60 countries with high-quality and reliable materials."},
    "nhu": {"specialty": "Personal care ingredients", "about": "NHU is a global leader in the manufacturing of vitamins, food additives, and cosmetic ingredients. Known for its strong R&D capabilities, NHU supplies essential nutrients and active ingredients to the global health and cosmetics sectors."},
    "nanogen": {"specialty": "Color cosmetics & film formers", "about": "Nanogen specializes in advanced technology for color cosmetics and film-forming polymers, providing performance ingredients that enhance visual appeal, wear-time, and texture in cosmetic formulations."},
    "mfci": {"specialty": "UV absorbers & sunscreen agents", "about": "MFCI is a premier global manufacturer of UV absorbers and organic UV filters for sunscreens and personal care products. They operate state-of-the-art production facilities committed to high standards of quality and regulatory compliance."},
    "sc-group": {"specialty": "Personal care ingredients", "about": "SpecChem (SC Group) is a leading developer and manufacturer of high-quality active ingredients for cosmetics and personal care, offering a comprehensive portfolio of skin-whitening, anti-aging, moisturizing, and soothing agents."},
    "tinci": {"specialty": "Personal care ingredients", "about": "Tinci is a major global producer of functional ingredients for personal care and cosmetics, specializing in surfactants, rheology modifiers, conditioning polymers, and eco-friendly raw materials."},
    "lingeba": {"specialty": "Personal care ingredients", "about": "Hangzhou Lingeba Technology is a specialty developer of cosmetic ingredients, personal care actives, and chemical formulations focused on delivering high performance and stability for modern cosmetic formulations."},
    "xinfu": {"specialty": "Personal care ingredients", "about": "Yifan Xinfu Pharmaceutical is a leading manufacturer of Vitamin B5 (D-Calcium Panthenate) and D-Panthenol. Serving both pharma, cosmetics, and food sectors, the company is globally recognized for its product purity and capacity."},
    "howtian": {"specialty": "Stevia extracts, inositol & personal care actives", "about": "HOWTIAN (formerly Zhucheng Haotian) is a global leader in natural sweetening solutions and plant-based ingredients. They are the world's leading manufacturer of premium Stevia extracts and key functional ingredients for food, beverages, and personal care."},
    "nexthia": {"specialty": "Personal care ingredients", "about": "Nexthia provides specialized chemical formulations and functional ingredients for personal care applications, helping cosmetic formulators optimize product stability, texture, and performance."},
    "3v-sigma": {"specialty": "Rheology modifiers, carbomers & conditioners", "about": "3V Sigma USA is a leading manufacturer of specialty chemicals including rheology modifiers, synthetic polymers, conditioning agents, and preservation systems for cosmetics, pharmaceuticals, and industrial applications."},
    "sabo": {"specialty": "Emollient esters & lipid enhancers", "about": "Sabo is a European chemical manufacturer specializing in emollient esters, emulsifiers, and lipid-enriching ingredients for cosmetics and personal care, prioritizing sustainable practices and quality formulations."},
    "akott": {"specialty": "Botanical actives & plant extracts", "about": "Akott is an Italian developer of premium active ingredients, botanical extracts, and cosmetic concepts. They merge nature with scientific research to provide high-efficacy solutions for anti-aging and skincare."},
    "alula": {"specialty": "Moringa Peregrina oils & extracts", "about": "Alula specializes in the extraction of premium Moringa Peregrina oil, a rare and highly nutritive botanical oil sourced from the Middle East, prized for its exceptional cosmetic, moisturizing, and anti-aging properties."},
    "salvona": {"specialty": "Encapsulation technology & microsphere delivery", "about": "Salvona Technologies is a pioneer in advanced encapsulation systems for personal care, creating sub-micron and microsphere delivery systems that enhance active ingredient stability, control release, and boost efficacy."},
    "zochem": {"specialty": "Personal care ingredients", "about": "Zochem is one of the largest producers of high-purity Zinc Oxide in North America, supplying critical ingredients for sunscreens, skincare, pharmaceuticals, and industrial products."},
    "olvea": {"specialty": "Natural oils, butters & argan oil", "about": "OLVEA is a leading French producer of natural vegetable and fish oils. Focused on sustainable sourcing, they supply high-quality organic vegetable oils, shea butter, and argan oil to the cosmetics and food industries."},
    "kiyu": {"specialty": "Personal care ingredients", "about": "Kiyu is a developer of specialized cosmetic actives, rheology modifiers, and emulsifiers, delivering functional raw materials that support modern formulations for skin and hair care."},
    "kci": {"specialty": "Personal care ingredients", "about": "Samyang KCI is a specialty chemical producer from South Korea, developing conditioning polymers, emulsifiers, and active raw materials that enhance texture, safety, and performance in personal care products."},
    "newman": {"specialty": "Personal care ingredients", "about": "Anhui Newman Fine Chemicals is a leading global manufacturer of Carbomer and polyacrylic acid polymers, providing high-quality thickeners, emulsifiers, and stabilizers for personal care and pharmaceutical formulations."},
    "yantai": {"specialty": "Personal care ingredients", "about": "Qingdao Aurora Chemical is a specialized developer of functional raw materials, active ingredients, and fine chemicals for the personal care, cosmetics, and feed sectors."},
    "xinyu": {"specialty": "Personal care ingredients", "about": "Jiangsu Xinyu Bio-Tech is a specialized manufacturer of chemical intermediates and cosmetic active ingredients, providing high-quality preservatives and antimicrobial agents to the personal care industry."},
    "cht": {"specialty": "Personal care ingredients", "about": "CHT is a global player in specialty chemicals and silicone-based formulations, supplying advanced emulsifiers, conditioning agents, and performance additives to the cosmetics and textile sectors."},
    "pellets": {"specialty": "Personal care ingredients", "about": "Chongqing Pellets Biotech is a specialized developer of encapsulated active beads, micro-capsules, and visual effect pigments that provide active delivery and sensory appeal in cosmetic formulations."},
    "lipidchem": {"specialty": "Fat & oil-based powders for pharma & cosmetics", "about": "Lipidchem SDN BHD is a Malaysian manufacturer of high-purity fatty acids, stearic acids, and lipid powders, supplying critical excipients and binders for pharmaceutical tableting and cosmetic formulations."},
    "orien": {"specialty": "Personal care fine chemicals & actives", "about": "QingZhou Orien Trading Co., Ltd. founded in 2016 with a complete technical service and product customization capability, is a professional enterprise engaged in the research and development of fine chemical products with more than 200 employees and a factory covering over 150 acres."},
    "satoria": {"specialty": "Resistant dextrin & pea protein", "about": "PT. Satoria Agro Industri is an Indonesian manufacturer specializing in the production of functional food ingredients, particularly plant-based and fiber-based solutions including Resistant Dextrin and Pea Protein."},
    "baltmilk": {"specialty": "Milk protein concentrate & micellar casein", "about": "BaltMilk is the cutting-edge manufacturer of milk proteins. We apply latest technology and innovation to provide the world with high value-added milk ingredients."},
    "tate-lyle": {"specialty": "FOS & GOS prebiotic fibers", "about": "Tate & Lyle is a global leader in sweetening, mouthfeel, and fiber fortification solutions, supplying high-quality prebiotic fibers like FOS and GOS, starches, and sugar substitutes for healthy nutrition."},
    "jianyuan": {"specialty": "Pea protein & plant-based proteins", "about": "JIANYUAN GROUP integrates production, processing and trade. Certified with ISO9001, HACCP, KOSHER, HALAL, BRC, Non-GMO, and ORGANIC. The main businesses are R&D, production and sale of pea, chickpea, mung bean, fava bean and other plant proteins, starches, dietary fibers, and medical protein."},
    "essential": {"specialty": "Flavorings & liquid essences", "about": "Essential Flavours has been bringing inspiring and on-trend flavours into Australian homes since 1989. With a dedication to liquid essences and powdered flavorings, they support major food and beverage creators globally."},
    "silvateam": {"specialty": "Pectins, tara gum & plant extracts", "about": "Silvateam is a global leader in plant-based extracts, manufacturing natural pectin, tara gum, and tannic acids for food texturizing, stabilizers, animal nutrition, and industrial processes."}
}

# Rule-based category classification for Pharma and Food
def classify_category(name, brand, industry, app=''):
    name_upper = name.upper()
    brand_upper = brand.upper() if brand else ''
    app_upper = app.upper() if app else ''
    combined = name_upper + ' ' + brand_upper + ' ' + app_upper
    
    if industry == 'cosmetics':
        if any(x in combined for x in ['UV FILTER', 'UV ABSORBER', 'SUNSCREEN', 'SUN SAFE', 'SUNSAFE', 'OCTOCRYLENE', 'TITANIUM DIOXIDE', 'BENZOPHENONE', 'OXYBENZONE', 'AVOBENZONE', 'HOMOSALATE', 'OCTISALATE', 'METHOXYCINNAMATE', 'ZINC OXIDE', 'SUN PROTECTION', 'PHOTO-PROTECTION', 'SPF']):
            return 'UV Filters'
        if any(x in combined for x in ['SURFACTANT', 'ISETHIONATE', 'COCOYL', 'LAUROYL', 'CLEANSING', 'BETAINE', 'SULFATE', 'SULPHATE', 'FOAMING']):
            return 'Surfactants'
        if any(x in combined for x in ['GLUCOSIDE']) and not any(x in combined for x in ['ASCORBYL GLUCOSIDE']):
            return 'Surfactants'
        if any(x in combined for x in ['CARBOMER', 'ACRYLATES', 'COPOLYMER', 'POLYMER', 'THICKENER', 'RHEOLOGY', 'CELLULOSE', 'XANTHAN', 'GUM', 'POLYQUATERNIUM', 'POLYQUAT', 'BEADS']):
            return 'Polymers'
        if any(x in combined for x in ['PRESERVATIVE', 'HYDROXYACETOPHENONE', 'PHENOXYETHANOL', 'PARABEN', 'ANTIMICROBIAL', 'ETHYLHEXYLGLYCERIN', 'ANTIFUNGAL', 'ANTIBACTERIAL', 'DANDRUFF']):
            return 'Preservatives'
        if any(x in combined for x in ['PIGMENT', 'IRON OXIDE', 'CARBON BLACK', 'BLACK 2', 'COLOR', 'MICA', 'DYE']):
            return 'Pigments'
        if any(x in combined for x in ['EMULSIFIER', 'EMULSION', 'STEARETH', 'CETEARETH', 'GLYCERYL STEARATE', 'POLYSORBATE', 'SORBITAN']):
            return 'Emulsifiers'
        if any(x in combined for x in ['EMOLLIENT', 'STEARIC ACID', 'TRIGLYCERIDE', 'OIL', 'BUTTER', 'LIPID', 'WAX', 'SQUALANE', 'DIMETHICONE', 'SILICONE', 'ARGAN', 'SHEA', 'MCT', 'ESTER']):
            return 'Emollients'
        # Default fallback for cosmetics is Actives since most functional cosmetic ingredients are actives
        return 'Actives'
        
    elif industry == 'pharma':
        if any(x in combined for x in ['HPMC', 'CELLULOSE', 'COAT', 'FILM', 'PVP', 'POVIDONE', 'COPOVIDONE', 'SHELLAC', 'PVA', 'POLYVINYL ALCOHOL', 'TITANIUM DIOXIDE', 'IRON OXIDE', 'SUCROSE', 'PEG', 'POLYETHYLENE GLYCOL', 'LOBILOT']):
            return 'Coatings'
        if any(x in combined for x in ['SORBITOL', 'MANNITOL', 'MALTITOL', 'DEXTROSE', 'TREHALOSE', 'XYLITOL', 'ERYTHRITOL', 'SUCRALOSE', 'SWEET', 'GLUCOSE']):
            return 'Sweeteners'
        if any(x in combined for x in ['ASCORBIC', 'ASCORBATE', 'VITAMIN', 'RIBOFLAVIN', 'PYRIDOXINE', 'THIAMINE', 'NIACIN', 'FOLIC', 'TOCOPHERYL', 'PANTOTHENATE', 'PANTHENOL']):
            return 'Vitamins'
        if any(x in combined for x in ['PHOSPHATE', 'CALCIUM', 'MAGNESIUM', 'ZINC', 'IRON', 'SULFATE', 'CARBONATE', 'OXIDE', 'FERROUS', 'SULPHATE', 'DICALCIUM']):
            return 'Minerals'
        return 'Excipients'
        
    elif industry == 'food':
        if any(x in combined for x in ['FIBER', 'BRAN', 'POMACE', 'DEXTRIN', 'POLYDEXTROSE', 'INULIN', 'PSYLLIUM', 'CELLULOSE', 'FLOUR']):
            return 'Dietary Fiber'
        if any(x in combined for x in ['PROTEIN', 'CASEIN', 'ISOLATE', 'CONCENTRATE', 'WHEY', 'GLUTEN']):
            return 'Proteins'
        if any(x in combined for x in ['GUM', 'XANTHAN', 'CARRAGEENAN', 'PECTIN', 'AGAR', 'ALGINATE', 'GUAR', 'LOCUST', 'STARCH', 'THICKENER']):
            return 'Thickeners'
        if any(x in combined for x in ['FOS', 'GOS', 'FRUCTO', 'GALACTO', 'OLIGOSACCHARIDE', 'PREBIOTIC']):
            return 'Prebiotics'
        if any(x in combined for x in ['LACTOSE', 'MILK', 'DAIRY', 'WPC', 'COLUSTRUM']):
            return 'Dairy Ingredients'
        if any(x in combined for x in ['ERYTHRITOL', 'XYLITOL', 'MALTITOL', 'SORBITOL', 'MANNITOL', 'STEVIA', 'SUCRALOSE', 'SWEETENER', 'MONK FRUIT']):
            return 'Sweeteners'
        return 'Specialty'
        
    return 'Specialty'

def to_js_obj(d):
    parts = []
    for k, v in d.items():
        if v is None:
            continue
        if isinstance(v, list):
            list_str = "[" + ", ".join(f'"{x}"' for x in v) + "]"
            parts.append(f'{k}: {list_str}')
        elif isinstance(v, str):
            escaped = v.replace('\\', '\\\\').replace('"', '\\"').replace('\n', '\\n').replace('\r', '')
            parts.append(f'{k}: "{escaped}"')
        else:
            parts.append(f'{k}: {v}')
    return "  { " + ", ".join(parts) + " }"

def main():
    excel_path = 'WEBSITE SCOPE 3 (1).xlsx'
    wb = openpyxl.load_workbook(excel_path)
    
    parsed_products = []
    partners_found = {}
    
    p_id_counter = 1
    
    # 1. Parse PHARMA
    pharma_sheet = wb['PHARMA']
    current_supplier = ''
    current_intro = ''
    for row_idx in range(3, pharma_sheet.max_row + 1):
        row = [cell.value for cell in pharma_sheet[row_idx]]
        if all(x is None for x in row):
            continue
        
        # Check repeated headers
        brand_val = row[3]
        prod_val = row[4]
        if not brand_val and not prod_val:
            continue
        if str(brand_val).strip().lower() == 'brand' or str(prod_val).strip().lower() == 'product':
            continue
            
        sup_val = row[1]
        if sup_val:
            current_supplier = str(sup_val).strip()
        intro_val = row[2]
        if intro_val:
            current_intro = str(intro_val).strip()
            
        partner_id = get_partner_id(current_supplier)
        partner_name = get_clean_partner_name(current_supplier, partner_id)
        
        # Save partner details
        if partner_id not in partners_found:
            partners_found[partner_id] = {
                'id': partner_id,
                'name': partner_name,
                'country': str(row[7]).strip() if row[7] else 'Global',
                'verticals': set(['pharma']),
                'about': current_intro
            }
        else:
            partners_found[partner_id]['verticals'].add('pharma')
            if current_intro and len(current_intro) > len(partners_found[partner_id]['about']):
                partners_found[partner_id]['about'] = current_intro
                
        # Create product
        prod_name = str(prod_val).strip()
        brand_name = str(brand_val).strip() if brand_val else ''
        grade_name = str(row[5]).strip() if row[5] else ''
        manuf = str(row[6]).strip() if row[6] else ''
        country = str(row[7]).strip() if row[7] else ''
        app = str(row[8]).strip() if row[8] else ''
        
        category = classify_category(prod_name, brand_name, 'pharma', app)
        
        product_obj = {
            'id': f'p{p_id_counter}',
            'name': prod_name,
            'brand': brand_name,
            'grade': grade_name,
            'principal': partner_name,
            'category': category,
            'industry': 'pharma',
            'description': app,
            'manufacturer': manuf,
            'country': country,
            'application': app
        }
        parsed_products.append(product_obj)
        p_id_counter += 1

    # 2. Parse PERSONAL CARE
    cosmetics_sheet = wb['PERSONAL CARE ']
    current_supplier = ''
    current_intro = ''
    for row_idx in range(2, cosmetics_sheet.max_row + 1):
        row = [cell.value for cell in cosmetics_sheet[row_idx]]
        if all(x is None for x in row):
            continue
        
        # Check repeated headers
        brand_val = row[3]
        inci_val = row[4]
        if not brand_val and not inci_val:
            continue
        if str(brand_val).strip().lower() == 'brand name' or str(inci_val).strip().lower() == 'inci name':
            continue
            
        sup_val = row[1]
        if sup_val:
            current_supplier = str(sup_val).strip()
        intro_val = row[2]
        if intro_val:
            current_intro = str(intro_val).strip()
            
        partner_id = get_partner_id(current_supplier)
        partner_name = get_clean_partner_name(current_supplier, partner_id)
        
        # Save partner details
        if partner_id not in partners_found:
            partners_found[partner_id] = {
                'id': partner_id,
                'name': partner_name,
                'country': str(row[6]).strip() if row[6] else 'Global',
                'verticals': set(['cosmetics']),
                'about': current_intro
            }
        else:
            partners_found[partner_id]['verticals'].add('cosmetics')
            if current_intro and len(current_intro) > len(partners_found[partner_id]['about']):
                partners_found[partner_id]['about'] = current_intro
                
        # Create product
        prod_inci = str(inci_val).strip()
        brand_name = str(brand_val).strip() if brand_val else ''
        manuf = str(row[5]).strip() if row[5] else ''
        country = str(row[6]).strip() if row[6] else ''
        app = str(row[7]).strip() if row[7] else ''
        
        category = classify_category(prod_inci, brand_name, 'cosmetics', app)
        
        product_obj = {
            'id': f'p{p_id_counter}',
            'name': prod_inci,
            'inci': prod_inci,
            'brand': brand_name,
            'principal': partner_name,
            'category': category,
            'industry': 'cosmetics',
            'description': app,
            'manufacturer': manuf,
            'country': country,
            'application': app
        }
        parsed_products.append(product_obj)
        p_id_counter += 1

    # 3. Parse FOOD
    food_sheet = wb['FOOD']
    current_supplier = ''
    current_intro = ''
    for row_idx in range(2, food_sheet.max_row + 1):
        row = [cell.value for cell in food_sheet[row_idx]]
        if all(x is None for x in row):
            continue
        
        # Check repeated headers
        brand_val = row[3]
        prod_val = row[4]
        if not brand_val and not prod_val:
            continue
        if str(brand_val).strip().lower() == 'brand' or str(prod_val).strip().lower() == 'product' or str(brand_val).strip().lower() == 'brand name' or str(prod_val).strip().lower() == 'product / grade':
            continue
            
        sup_val = row[1]
        if sup_val:
            current_supplier = str(sup_val).strip()
        intro_val = row[2]
        if intro_val:
            current_intro = str(intro_val).strip()
            
        partner_id = get_partner_id(current_supplier)
        partner_name = get_clean_partner_name(current_supplier, partner_id)
        
        # Save partner details
        if partner_id not in partners_found:
            partners_found[partner_id] = {
                'id': partner_id,
                'name': partner_name,
                'country': str(row[7]).strip() if row[7] else 'Global',
                'verticals': set(['food']),
                'about': current_intro
            }
        else:
            partners_found[partner_id]['verticals'].add('food')
            if current_intro and len(current_intro) > len(partners_found[partner_id]['about']):
                partners_found[partner_id]['about'] = current_intro
                
        # Create product
        prod_name = str(prod_val).strip()
        brand_name = str(brand_val).strip() if brand_val else ''
        grade_name = str(row[5]).strip() if row[5] else ''
        manuf = str(row[6]).strip() if row[6] else ''
        country = str(row[7]).strip() if row[7] else ''
        app = str(row[8]).strip() if row[8] else ''
        
        category = classify_category(prod_name, brand_name, 'food', app)
        
        product_obj = {
            'id': f'p{p_id_counter}',
            'name': prod_name,
            'brand': brand_name,
            'grade': grade_name,
            'principal': partner_name,
            'category': category,
            'industry': 'food',
            'description': app,
            'manufacturer': manuf,
            'country': country,
            'application': app
        }
        parsed_products.append(product_obj)
        p_id_counter += 1

    # Find existing logos
    existing_logos = os.listdir('public/logos')
    def find_partner_logo(pid):
        # Try exact lowercase matches
        for ext in ['.png', '.jpg', '.jpeg', '.svg']:
            fn = f"{pid}{ext}"
            if fn in existing_logos:
                return fn
        
        # Fallback string matching
        for f in existing_logos:
            if f.lower().startswith(pid) and f.lower().endswith(('.png', '.jpg', '.jpeg', '.svg')):
                return f
        
        # Special check for SC Group / Newman / howtian sopure
        if pid == 'sc-group':
            for f in ['sc-group.png', '斯拜科LOGOPNG.PNG']:
                if f in existing_logos: return f
        elif pid == 'howtian':
            for f in ['howtian.png', 'sopure.jpg', 'SOPURE LOGO.JPG']:
                if f in existing_logos: return f
        elif pid == 'yantai':
            for f in ['aurora-chem.jpg']:
                if f in existing_logos: return f
        elif pid == 'silvateam':
            for f in ['jrs-silvateam.png', 'Silva team - logo.png']:
                if f in existing_logos: return f
        elif pid == 'essential':
            for f in ['essential.png', 'Essential logo.png']:
                if f in existing_logos: return f
        elif pid == 'orien':
            for f in ['ailitong.png']:
                if f in existing_logos: return f
        elif pid == 'satoria':
            for f in ['satoria.svg']:
                if f in existing_logos: return f
                
        return None

    # Merge partner details
    partners_list = []
    for pid, pdata in partners_found.items():
        # Fallbacks or existing hand-crafted metadata
        merged = {
            'id': pid,
            'name': pdata['name'],
            'country': pdata['country'] if pdata['country'] else 'Global',
            'verticals': sorted(list(pdata['verticals'])),
            'specialty': '',
            'about': pdata['about']
        }
        
        logo_file = find_partner_logo(pid)
        if logo_file:
            merged['logo'] = logo_file
            
        # Merge existing handcrafted details
        if pid in existing_partners:
            ext = existing_partners[pid]
            if ext.get('specialty'):
                merged['specialty'] = ext['specialty']
            if ext.get('about') and (not merged['about'] or len(merged['about']) < 30):
                merged['about'] = ext['about']
                
        if not merged['specialty']:
            # Assign a default specialty based on verticals
            v = merged['verticals'][0]
            if v == 'pharma':
                merged['specialty'] = 'Pharmaceutical excipients supplier'
            elif v == 'cosmetics':
                merged['specialty'] = 'Personal care ingredients supplier'
            else:
                merged['specialty'] = 'Food ingredients supplier'
                
        partners_list.append(merged)
        
    # Sort partners by name
    partners_list.sort(key=lambda x: x['name'])
    
    # 4. Generate partners.ts
    partners_content = """export interface Partner {
  id: string;
  name: string;
  country: string;
  verticals: ("pharma" | "cosmetics" | "food")[];
  specialty: string;
  logo?: string;
  about?: string;
}

export const partners: Partner[] = [
"""
    for partner in partners_list:
        partners_content += to_js_obj(partner) + ",\n"
    partners_content += """];

export const getPartnersByVertical = (vertical: "pharma" | "cosmetics" | "food") =>
  partners.filter((p) => p.verticals.includes(vertical));
"""
    
    with open('src/data/partners.ts', 'w') as f:
        f.write(partners_content)
    print('Generated src/data/partners.ts')
    
    # 5. Generate products.ts
    products_content = """export interface Product {
  id: string;
  name: string;
  inci?: string;
  brand?: string;
  grade?: string;
  principal: string;
  category: string;
  industry: "pharma" | "cosmetics" | "food";
  description: string;
  form?: string;
  manufacturer?: string;
  country?: string;
  application?: string;
}

export const products: Product[] = [
"""
    for prod in parsed_products:
        products_content += to_js_obj(prod) + ",\n"
    products_content += """];

export const categories = [...new Set(products.map((p) => p.category))];
export const industries = ["pharma", "cosmetics", "food"] as const;

export const getProductsByIndustry = (industry: string) =>
  products.filter((p) => p.industry === industry);

export const getProductsByCategory = (category: string) =>
  products.filter((p) => p.category === category);
"""
    
    with open('src/data/products.ts', 'w') as f:
        f.write(products_content)
    print(f'Generated src/data/products.ts with {len(parsed_products)} products')

if __name__ == '__main__':
    main()
